"""
Read Excel (.xlsx, .xls) into list of dicts (first data row = headers).
Used by load_drg, load_apc, load_asp_pricing for CMS files.
"""
from pathlib import Path


def _normalize_header(cell_value) -> str:
    """Strip leading/trailing whitespace from header (e.g. 'APC ' -> 'APC')."""
    if cell_value is None:
        return ""
    return str(cell_value).strip()


def find_header_row(path: Path, sheet_index: int = 0, sentinel: str = "HCPCS Code", max_search: int = 200) -> int | None:
    """
    Find the 0-based row index where the first cell (after strip) equals sentinel.
    Scans from row 0 up to max_search. Returns None if not found.
    """
    path = Path(path)
    suffix = path.suffix.lower()
    if suffix == ".xlsx":
        return _find_header_row_xlsx(path, sheet_index, sentinel, max_search)
    if suffix == ".xls":
        return _find_header_row_xls(path, sheet_index, sentinel, max_search)
    return None


def _find_header_row_xlsx(path: Path, sheet_index: int, sentinel: str, max_search: int) -> int | None:
    import openpyxl
    sentinel = sentinel.strip()
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = wb.worksheets[sheet_index]
        for row_idx, row in enumerate(sheet.iter_rows(values_only=True)):
            if row_idx >= max_search:
                break
            if not row:
                continue
            first_cell = row[0] if isinstance(row, (list, tuple)) else row
            if first_cell is not None and str(first_cell).strip() == sentinel:
                return row_idx
    finally:
        wb.close()
    return None


def _find_header_row_xls(path: Path, sheet_index: int, sentinel: str, max_search: int) -> int | None:
    import xlrd
    sentinel = sentinel.strip()
    with xlrd.open_workbook(path) as wb:
        sheet = wb.sheet_by_index(sheet_index)
        for r in range(min(max_search, sheet.nrows)):
            try:
                first_cell = sheet.cell_value(r, 0)
            except IndexError:
                continue
            if first_cell is not None and str(first_cell).strip() == sentinel:
                return r
    return None


def read_excel_to_dicts(
    path: Path,
    skip_rows: int = 0,
    sheet_index: int = 0,
    find_header_sentinel: str | None = None,
    max_header_search: int = 200,
):
    """
    Read sheet of an Excel file into a list of dicts.
    Row at skip_rows is treated as header; following rows as data.
    All column headers are stripped (leading/trailing whitespace) before use as keys.
    If find_header_sentinel is set (e.g. "HCPCS Code"), the first row whose first
    cell equals that string (after strip) is used as the header row; skip_rows is ignored.
    """
    path = Path(path)
    suffix = path.suffix.lower()
    if find_header_sentinel and suffix in (".xlsx", ".xls"):
        found = find_header_row(path, sheet_index, find_header_sentinel.strip(), max_header_search)
        if found is not None:
            skip_rows = found
    if suffix == ".xlsx":
        return _read_xlsx(path, skip_rows, sheet_index)
    if suffix == ".xls":
        return _read_xls(path, skip_rows, sheet_index)
    raise ValueError(f"Unsupported Excel extension: {suffix}")


def _read_xlsx(path: Path, skip_rows: int, sheet_index: int):
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = wb.worksheets[sheet_index]
        rows = list(sheet.iter_rows(values_only=True))
    finally:
        wb.close()
    if not rows or len(rows) <= skip_rows:
        return []
    header = [_normalize_header(c) for c in rows[skip_rows]]
    out = []
    for row in rows[skip_rows + 1 :]:
        if row is None:
            continue
        cells = list(row) if isinstance(row, (list, tuple)) else [row]
        while len(cells) < len(header):
            cells.append(None)
        d = {}
        for i, key in enumerate(header):
            if key:
                val = cells[i] if i < len(cells) else None
                d[key] = val
        if d and any(v is not None and str(v).strip() for v in d.values()):
            out.append(d)
    return out


def _read_xls(path: Path, skip_rows: int, sheet_index: int):
    import xlrd
    with xlrd.open_workbook(path) as wb:
        sheet = wb.sheet_by_index(sheet_index)
        if sheet.nrows <= skip_rows:
            return []
        header = [_normalize_header(sheet.cell_value(skip_rows, c)) for c in range(sheet.ncols)]
        out = []
        for r in range(skip_rows + 1, sheet.nrows):
            d = {}
            for c, key in enumerate(header):
                if key:
                    try:
                        val = sheet.cell_value(r, c)
                    except IndexError:
                        val = None
                    d[key] = val
            if d and any(v is not None and str(v).strip() for v in d.values()):
                out.append(d)
        return out


def pick_column(row_dict: dict, *alias_keys: str):
    """Return first non-empty value from row for any of the given header names (case-insensitive)."""
    keys_lower = {str(k).lower().strip(): k for k in row_dict if k is not None and str(k).strip()}
    for want in alias_keys:
        w = str(want).lower().strip()
        for orig, key in keys_lower.items():
            if w in orig or orig.startswith(w) or w.startswith(orig):
                val = row_dict.get(key)
                if val is not None and str(val).strip() != "":
                    return val
    for want in alias_keys:
        if want in row_dict:
            val = row_dict[want]
            if val is not None and str(val).strip() != "":
                return val
    return None


def map_row(row_dict: dict, field_aliases: dict) -> dict:
    """
    Map a row dict (Excel headers as keys) to a dict with our field names.
    field_aliases: { 'our_key': ['CMS header 1', 'CMS header 2', ...], ... }
    """
    out = {}
    for our_key, aliases in field_aliases.items():
        v = pick_column(row_dict, *aliases)
        out[our_key] = v
    return out
